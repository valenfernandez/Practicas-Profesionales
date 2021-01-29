import openpyxl


def excel_to_set(doc, first_cell, last_cell):
    """
    Esta funcion devuelve una lista con todas las frases unicas almacenadas en un grupo de celdas de un excel
    doc: string nombre del archivo excel a trabajar.
    first_cell: primera celda del rango a copiar.
    last_cell: ultima celda del rango a copiar
    """
    phrases = []
    excel_document = openpyxl.load_workbook(doc, data_only=True)
    first_sheet = excel_document.sheetnames[0]
    sheet = excel_document[first_sheet]
    for row in sheet[first_cell:last_cell]:
        for cell in row:
            if cell.value and cell.value != 'no_corresponde':  # si la celda tiene informacion
                cell_with_spaces = cell.value.replace('_', ' ')
                phrases.append(cell_with_spaces)
    return set(phrases)


def excel_to_text(doc, first_cell, last_cell):
    """ Esta funcion devuelve en forma de string los datos relevantes almacenados en un grupo de celdas de un excel
        doc: string nombre del archivo excel a trabajar.
        first_cell: primera celda del rango a copiar.
        last_cell: ultima celda del rango a copiar
        """
    phrases = excel_to_set(doc, first_cell, last_cell)
    text = ""
    for phrase in phrases:
        text += f'{phrase}\n'
    return text


def excel_to_data(doc, first_cell_num, last_cell_num, col_phrases, col_cats):
    """ Esta funcion devuelve una lista con todas las frases unicas encontradas en un rango de celdas de un excel
    y una lista con las categorias amenazante, no amenazante para cada una de estas frases.
    doc: string nombre del archivo excel a trabajar.
    first_cell_num: numero de la primera celda del rango a copiar.
    last_cell_num:  numero de la ultima celda del rango a copiar
    col_phrases: columna donde estan las frases en el archivo doc
    col_cats: columna donde esta la clasificacion en el archivo doc
     """
    cat_list = []
    text_list = []
    excel_document = openpyxl.load_workbook(doc, data_only=True)
    first_sheet = excel_document.sheetnames[0]
    sheet = excel_document[first_sheet]
    for i in range(first_cell_num, last_cell_num):
        cell = sheet[col_phrases + str(i)]
        if cell.value and (cell.value != 'no corresponde') :
            text = cell.value.replace('_', ' ')
            if (text not in text_list):
                cat_cell = sheet[col_cats + str(i)]
                if cat_cell.value == 'amenazas':
                    cats = {'AMENAZANTE': 1, 'NO_AMENAZANTE': 0}
                else:
                    cats = {'AMENAZANTE': 0, 'NO_AMENAZANTE': 1}
                cat_list.append(cats)
                text_list.append(text)
    return text_list, cat_list


def num_relevant_cells(doc, first_cell, last_cell):
    """Esta funcion escribe por pantalla estadisticas de el numero de celdas
    con informacion no nula en un rango de celdas de una columna en un excel

    doc: string nombre del archivo excel a trabajar.
    first_cell: primera celda del rango
    last_cell: ultima celda del rango
    """
    percentage = 0
    relevant_cells = len(excel_to_set(doc, first_cell, last_cell))
    total = int(last_cell[1:]) - int(first_cell[1:])
    if total != 0:
        percentage = relevant_cells * 100 / total
    print(f'el numero total de celdas: {total} \n'
          f'el numero total de celdas con informacion unica: {relevant_cells}\n'
          f'eso es un %{percentage:.2f} de celdas con informacion')
